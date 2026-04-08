from __future__ import annotations

import json
from dataclasses import asdict, is_dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from excel_diff.models import ComparisonResult


def write_outputs(result: ComparisonResult, output_dir: str | Path | None = None, output_name: str | None = None) -> tuple[Path, Path]:
    target_dir = Path(output_dir or Path.cwd() / "saida")
    target_dir.mkdir(parents=True, exist_ok=True)
    stem = output_name or f"diff_{result.base_profile.sheet_name}_vs_{result.compare_profile.sheet_name}"
    excel_path = target_dir / f"{stem}.xlsx"
    json_path = target_dir / f"{stem}.json"

    write_excel_report(result, excel_path)
    write_json_report(result, json_path)
    return excel_path, json_path


def write_excel_report(result: ComparisonResult, path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"

    bold = Font(bold=True)
    summary_rows = [
        ("Arquivo base", result.base_profile.path),
        ("Aba base", result.base_profile.sheet_name),
        ("Arquivo comparação", result.compare_profile.path),
        ("Aba comparação", result.compare_profile.sheet_name),
        ("Chave base", result.key_column),
        ("Chave comparação", result.resolved_compare_key),
        ("Linhas somente base", len(result.only_in_base)),
        ("Linhas somente comparação", len(result.only_in_compare)),
        ("Linhas comparadas", len(result.matched_rows)),
        ("Problemas de validação", len(result.validation_issues)),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=row_index, column=1, value=label).font = bold
        summary_sheet.cell(row=row_index, column=2, value=value)

    if result.validation_issues:
        start_row = len(summary_rows) + 3
        summary_sheet.cell(row=start_row, column=1, value="Validações").font = bold
        for offset, issue in enumerate(result.validation_issues, start=1):
            summary_sheet.cell(row=start_row + offset, column=1, value=issue.level)
            summary_sheet.cell(row=start_row + offset, column=2, value=issue.code)
            summary_sheet.cell(row=start_row + offset, column=3, value=issue.message)

    write_column_mappings(workbook, result)
    write_differences_sheet(workbook, "Diferenças", result.matched_rows)
    write_differences_sheet(workbook, "Somente_Base", result.only_in_base)
    write_differences_sheet(workbook, "Somente_Comparacao", result.only_in_compare)

    workbook.save(path)


def write_column_mappings(workbook: Workbook, result: ComparisonResult) -> None:
    sheet = workbook.create_sheet("Mapeamento")
    headers = ["Coluna Base", "Coluna Comparacao", "Metodo", "Score"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for mapping in result.column_mappings:
        sheet.append([mapping.base_column, mapping.compare_column, mapping.method, mapping.score])


def write_differences_sheet(workbook: Workbook, name: str, rows: list) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(["Chave", "Linha Base", "Linha Comparacao", "Status", "Mudanças"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.key, row.base_row_number, row.compare_row_number, row.status, json.dumps(row.changes, ensure_ascii=False)])


def write_json_report(result: ComparisonResult, path: Path) -> None:
    payload = dataclass_to_dict(result)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def dataclass_to_dict(value):
    if is_dataclass(value):
        return {key: dataclass_to_dict(item) for key, item in asdict(value).items()}
    if isinstance(value, list):
        return [dataclass_to_dict(item) for item in value]
    if isinstance(value, dict):
        return {key: dataclass_to_dict(item) for key, item in value.items()}
    return value
