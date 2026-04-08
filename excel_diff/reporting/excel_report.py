from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from excel_diff.models import ComparisonResult


def write_excel_report(result: ComparisonResult, path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"

    bold = Font(bold=True)
    summary_rows = [
        ("Arquivo base", result.base_profile.path),
        ("Aba base", result.base_profile.sheet_name),
        ("Arquivo comparação", result.compare_profile.path),
        ("Aba comparação", result.compare_profile.sheet_name),
        ("Chave base", result.key_column),
        ("Chave comparação", result.resolved_compare_key),
        ("Chaves de identificação base", ", ".join(pair.base_column for pair in result.diff_key_pairs)),
        ("Chaves de identificação comparação", ", ".join(pair.compare_column for pair in result.diff_key_pairs)),
        ("Linhas somente base", len(result.only_in_base)),
        ("Linhas somente comparação", len(result.only_in_compare)),
        ("Linhas alteradas", len([row for row in result.matched_rows if row.status == "changed"])),
        ("Linhas iguais", len([row for row in result.matched_rows if row.status == "matched"])),
        ("Problemas de validação", len(result.validation_issues)),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=row_index, column=1, value=label).font = bold
        summary_sheet.cell(row=row_index, column=2, value=value)

    if result.validation_issues:
        start_row = len(summary_rows) + 3
        summary_sheet.cell(row=start_row, column=1, value="Validações").font = bold
        for offset, issue in enumerate(result.validation_issues, start=1):
            summary_sheet.cell(row=start_row + offset, column=1, value=issue.level)
            summary_sheet.cell(row=start_row + offset, column=2, value=issue.code)
            summary_sheet.cell(row=start_row + offset, column=3, value=issue.message)

    write_column_mappings(workbook, result)
    write_diff_sheet(workbook, "Adição", result.only_in_compare, result)
    write_diff_sheet(workbook, "Exclusão", result.only_in_base, result)
    write_diff_sheet(workbook, "Alteração", [row for row in result.matched_rows if row.status == "changed"], result)
    write_diff_sheet(workbook, "Igual", [row for row in result.matched_rows if row.status == "matched"], result)

    workbook.save(path)


def write_column_mappings(workbook: Workbook, result: ComparisonResult) -> None:
    sheet = workbook.create_sheet("Mapeamento")
    headers = ["Coluna Base", "Coluna Comparacao", "Metodo", "Score"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for mapping in result.column_mappings:
        sheet.append([mapping.base_column, mapping.compare_column, mapping.method, mapping.score])


def write_diff_sheet(workbook: Workbook, sheet_name: str, rows: list, result: ComparisonResult) -> None:
    sheet = workbook.create_sheet(sheet_name)
    headers = ["Categoria", "Chave", "Linha Base", "Linha Comparacao", "Status"]
    for pair in result.diff_key_pairs:
        headers.extend([
            f"Base - {pair.base_column}",
            f"Comparacao - {pair.compare_column}",
            f"Diff - {pair.base_column} x {pair.compare_column}",
        ])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        values = [sheet_name, row.key, row.base_row_number, row.compare_row_number, row.status]
        for identifier in row.diff_identifiers:
            base_value = identifier.get("base_value")
            compare_value = identifier.get("compare_value")
            values.extend([
                base_value,
                compare_value,
                diff_cell_value(base_value, compare_value),
            ])
        sheet.append(values)


def diff_cell_value(base_value, compare_value) -> str:
    if base_value == compare_value:
        return "OK"
    return f"{base_value} -> {compare_value}"