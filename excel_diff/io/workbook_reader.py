from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def open_workbook(path: str | Path):
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {workbook_path}")
    return load_workbook(workbook_path, read_only=True, data_only=True)


def list_sheets(path: str | Path) -> list[str]:
    workbook = open_workbook(path)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def iter_sheet_rows(path: str | Path, sheet_name: str, start_row: int = 1, max_rows: int | None = None) -> Iterable[tuple]:
    workbook = open_workbook(path)
    try:
        sheet = workbook[sheet_name]
        if max_rows is None:
            rows = sheet.iter_rows(min_row=start_row, values_only=True)
        else:
            end_row = start_row + max_rows - 1
            rows = sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True)
        for row in rows:
            yield row
    finally:
        workbook.close()


def sample_sheet_rows(path: str | Path, sheet_name: str, max_rows: int = 100) -> list[tuple]:
    return list(iter_sheet_rows(path, sheet_name, start_row=1, max_rows=max_rows))


def read_sheet_data(path: str | Path, sheet_name: str, start_row: int) -> list[tuple]:
    workbook = open_workbook(path)
    try:
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, values_only=True):
            rows.append(row)
        return rows
    finally:
        workbook.close()
